# mrg.py
import streamlit as st
import pandas as pd
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import base64  # For base64 image encoding

# Check for openpyxl availability
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    st.error(
        "The 'openpyxl' library is not installed. Please ensure it's included in your requirements.txt and the environment is set up correctly.")
    st.stop()


def apply_main_sheet_conditional_formatting(ws, sheet_name_in_wb, workbook_obj, low_thresh, mid_thresh):
    temp_buffer_for_df = io.BytesIO()
    workbook_obj.save(temp_buffer_for_df) # Save current state of workbook to read from
    temp_buffer_for_df.seek(0)
    try:
        df = pd.read_excel(temp_buffer_for_df, sheet_name=sheet_name_in_wb)
        if df.empty: 
            return
    except ValueError as e:
        st.warning(f"Could not read sheet '{sheet_name_in_wb}' for formatting. Skipping. Error: {e}")
        return

    dark_green_fill_main = PatternFill(start_color='19D119', end_color='19D119', fill_type='solid')
    dark_red_fill_main = PatternFill(start_color='E82D1C', end_color='E82D1C', fill_type='solid')
    
    presence_col_name = 'presence'
    presence_col_df_idx = df.columns.get_loc(presence_col_name) if presence_col_name in df.columns else None

    for df_col_idx, df_col_name in enumerate(df.columns):
        excel_col_letter = get_column_letter(df_col_idx + 1)
        for excel_row_num in range(2, len(df) + 2): 
            df_row_idx = excel_row_num - 2
            cell = ws[f'{excel_col_letter}{excel_row_num}']
            # Ensure value is read from DataFrame, as cell.value might be None if sheet was just created
            if df_row_idx < len(df) and df_col_idx < len(df.columns):
                 value = df.iat[df_row_idx, df_col_idx]
            else:
                continue # Should not happen if df is read correctly

            if df_col_name.endswith('_Diff'):
                if pd.notna(value) and isinstance(value, (int, float)):
                    cell.value = value 
                    cell.number_format = '0.00%'
                    if value <= low_thresh: cell.fill = dark_green_fill_main
                    elif value <= mid_thresh:
                        if mid_thresh > low_thresh:
                            ratio = (value - low_thresh) / (mid_thresh - low_thresh)
                            r_comp = max(0, min(int(255 + (139 - 255) * ratio), 255))
                            g_comp = max(0, min(int(255 - (255 - 0) * ratio), 255))
                            b_comp = 0
                            color_hex = f'{r_comp:02X}{g_comp:02X}{b_comp:02X}'
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        else: cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
                    else: cell.fill = dark_red_fill_main
            
            elif presence_col_df_idx is not None and df_col_idx == presence_col_df_idx:
                if str(value) == 'Present in Both': cell.fill = dark_green_fill_main
                elif str(value) in ['Present in excel', 'Present in PBI']: cell.fill = dark_red_fill_main
    
    if not df.empty:
        summary_row_blue = '80BBD9'
        summary_fill = PatternFill(start_color=summary_row_blue, end_color=summary_row_blue, fill_type='solid')
        header_row_darker_blue = '609AB9'
        header_fill = PatternFill(start_color=header_row_darker_blue, end_color=header_row_darker_blue, fill_type='solid')
        bold_font = Font(bold=True)

        for col_excel_idx in range(1, ws.max_column + 1): 
            ws.cell(row=1, column=col_excel_idx).font = bold_font
            ws.cell(row=1, column=col_excel_idx).fill = header_fill
        
        if ws.max_row >= 2: 
            for col_excel_idx in range(1, ws.max_column + 1):
                ws.cell(row=2, column=col_excel_idx).font = bold_font
                ws.cell(row=2, column=col_excel_idx).fill = summary_fill


def combine_excel_files(file_list, low_threshold, mid_threshold):
    if not file_list or len(file_list) > 10:
        st.error("Please upload 1 to 10 files.")
        return None, None

    first_filename_parts = os.path.splitext(file_list[0].name)[0].split('_')
    base_name = first_filename_parts[0] if first_filename_parts else os.path.splitext(file_list[0].name)[0]
    output_filename = f"{base_name}_MERGED_reports.xlsx"

    output_buffer = io.BytesIO()
    output_wb = Workbook()
    if 'Sheet' in output_wb.sheetnames: output_wb.remove(output_wb['Sheet'])

    data_sheet_names_in_output = []
    # sheet_name_output_counts tracks occurrences of *original_sheet_name* to generate initial suffixes
    sheet_name_output_counts = {} 
    all_pages_summary_data = []
    total_avg_diff_sum_for_pooled = 0.0
    num_sheets_for_pooled_avg = 0

    for uploaded_file in file_list:
        file_bytes = uploaded_file.read()
        try:
            current_input_wb = load_workbook(filename=io.BytesIO(file_bytes))
        except Exception as e:
            st.warning(f"Could not read {uploaded_file.name}: {e}. Skipping this file.")
            continue

        for original_sheet_name in current_input_wb.sheetnames:
            if original_sheet_name in ["Column_Checklist", "Diff_Checker_Summary", "All_Pages_Summary"]:
                continue

            ws_source = current_input_wb[original_sheet_name]
            
            # --- Refined Sheet Naming Logic ---
            occurrence_count = sheet_name_output_counts.get(original_sheet_name, 0)
            sheet_name_output_counts[original_sheet_name] = occurrence_count + 1

            candidate_sheet_name = original_sheet_name
            if occurrence_count > 0: # Not the first time we've seen this original_sheet_name
                suffix = f"_{occurrence_count}"
                # Try to keep original name + suffix, then truncate
                candidate_sheet_name = f"{original_sheet_name}{suffix}"
            
            # Truncate the candidate name if it's too long
            truncated_candidate_name = candidate_sheet_name[:31]

            # Ensure uniqueness of the (potentially truncated) name in the output workbook
            final_target_sheet_name = truncated_candidate_name
            clash_resolution_counter = 0
            while final_target_sheet_name in output_wb.sheetnames:
                clash_resolution_counter += 1
                # If truncated_candidate_name was already 31 chars, we need to shorten it to add suffix
                base_for_clash_suffix = truncated_candidate_name
                suffix_for_clash = f"({clash_resolution_counter})"
                
                if len(base_for_clash_suffix) + len(suffix_for_clash) > 31:
                    base_for_clash_suffix = base_for_clash_suffix[:31 - len(suffix_for_clash)]
                
                final_target_sheet_name = f"{base_for_clash_suffix}{suffix_for_clash}"
                if clash_resolution_counter > 50: # Safety break
                    st.error(f"Extreme difficulty generating unique name for {original_sheet_name}")
                    final_target_sheet_name = f"ERR_NAME_{len(output_wb.sheetnames)}"[:31] # Fallback
                    break
            # --- End of Refined Sheet Naming Logic ---
            
            data_sheet_names_in_output.append(final_target_sheet_name)
            ws_target = output_wb.create_sheet(title=final_target_sheet_name)
            for row in ws_source.rows: 
                for cell in row: ws_target[cell.coordinate].value = cell.value
            
            avg_diff_display_text = "N/A"
            presence_display_text = "N/A"
            avg_diff_numeric = None

            if ws_source.max_row >= 2:
                a2_val = ws_source.cell(row=2, column=1).value
                if a2_val and isinstance(a2_val, str) and "Avg Diff:" in a2_val:
                    avg_diff_display_text = a2_val 
                    try:
                        perc_str = avg_diff_display_text.split("Avg Diff:")[1].strip().replace('%', '')
                        avg_diff_numeric = float(perc_str) / 100.0
                        total_avg_diff_sum_for_pooled += avg_diff_numeric
                        num_sheets_for_pooled_avg += 1
                    except (IndexError, ValueError): avg_diff_numeric = None
                elif pd.notna(a2_val) : avg_diff_display_text = str(a2_val)

                presence_col_idx_src = None
                for col_scan in range(1, ws_source.max_column + 1):
                    if ws_source.cell(row=1, column=col_scan).value == 'presence':
                        presence_col_idx_src = col_scan; break
                if presence_col_idx_src: presence_display_text = str(ws_source.cell(row=2, column=presence_col_idx_src).value or "N/A")
            
            display_name_for_summary = final_target_sheet_name # Use the actual unique name in output
            suffixes_to_remove = ["_validation_report", "_val_report", "_validationreport", "_val"] 
            # For display in summary, try to clean it further
            cleaned_display_name = final_target_sheet_name 
            for suffix in suffixes_to_remove:
                if cleaned_display_name.lower().endswith(suffix.lower()): 
                    cleaned_display_name = cleaned_display_name[:-len(suffix)]
                    break 
            if not cleaned_display_name.strip(): cleaned_display_name = final_target_sheet_name

            all_pages_summary_data.append({
                'Display Sheet Name': cleaned_display_name, # Use the further cleaned name for display
                'Actual Sheet Name': final_target_sheet_name, # Keep track of actual name if needed
                'Presence': presence_display_text,
                'Avg Diff Numeric': avg_diff_numeric,
                'Avg Diff Original Text': avg_diff_display_text 
            })

    for sheet_name_to_fmt in data_sheet_names_in_output:
        if sheet_name_to_fmt in output_wb.sheetnames:
            apply_main_sheet_conditional_formatting(output_wb[sheet_name_to_fmt], sheet_name_to_fmt, output_wb, low_threshold, mid_threshold)

    summary_page_title = "All_Pages_Summary"
    if summary_page_title in output_wb.sheetnames: del output_wb[summary_page_title]
    summary_ws = output_wb.create_sheet(title=summary_page_title, index=0)
    
    headers = ["Sheet Name", "Presence", "Avg Diff"]
    for col_num, header_text in enumerate(headers, 1):
        summary_ws.cell(row=1, column=col_num, value=header_text).font = Font(bold=True)
    summary_ws.column_dimensions['A'].width = 35
    summary_ws.column_dimensions['B'].width = 45
    summary_ws.column_dimensions['C'].width = 20

    dark_green_fill_summary = PatternFill(start_color='19D119', end_color='19D119', fill_type='solid')
    dark_red_fill_summary = PatternFill(start_color='E82D1C', end_color='E82D1C', fill_type='solid')

    summary_row_idx = 2
    for item in all_pages_summary_data:
        summary_ws.cell(row=summary_row_idx, column=1, value=item['Display Sheet Name']) # Show cleaned name
        summary_ws.cell(row=summary_row_idx, column=2, value=item['Presence'])
        
        avg_diff_val_numeric = item['Avg Diff Numeric']
        cell_c_summary = summary_ws.cell(row=summary_row_idx, column=3)

        if avg_diff_val_numeric is not None and isinstance(avg_diff_val_numeric, (float, int)):
            cell_c_summary.value = avg_diff_val_numeric
            cell_c_summary.number_format = '0.00%'
            if avg_diff_val_numeric <= low_threshold: cell_c_summary.fill = dark_green_fill_summary 
            elif avg_diff_val_numeric <= mid_threshold:
                if mid_threshold > low_threshold:
                    ratio = (avg_diff_val_numeric - low_threshold) / (mid_threshold - low_threshold)
                    r_comp = max(0, min(int(255 + (139 - 255) * ratio), 255))
                    g_comp = max(0, min(int(255 - (255 - 0) * ratio), 255))
                    b_comp = 0
                    color_hex = f'{r_comp:02X}{g_comp:02X}{b_comp:02X}'
                    cell_c_summary.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                else: cell_c_summary.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            else: cell_c_summary.fill = dark_red_fill_summary
        else: 
             cell_c_summary.value = item['Avg Diff Original Text'] 
        summary_row_idx += 1
    
    summary_ws.cell(row=summary_row_idx, column=1, value="Pooled Average").font = Font(bold=True)
    if num_sheets_for_pooled_avg > 0:
        pooled_avg = total_avg_diff_sum_for_pooled / num_sheets_for_pooled_avg
        cell_pooled_c = summary_ws.cell(row=summary_row_idx, column=3, value=pooled_avg)
        cell_pooled_c.number_format = '0.00%'
        cell_pooled_c.font = Font(bold=True)
    else:
        summary_ws.cell(row=summary_row_idx, column=3, value="N/A").font = Font(bold=True)

    final_ordered_sheet_names = [summary_page_title] + [name for name in data_sheet_names_in_output if name != summary_page_title]
    if final_ordered_sheet_names: 
        output_wb._sheets = [output_wb[name] for name in final_ordered_sheet_names if name in output_wb.sheetnames]

    output_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer, output_filename


def run():
    st.markdown("""
        <style>
        .title { font-size: 36px; color: #FF4B4B; text-align: center; font-weight: bold; margin-bottom: 20px; }
        .instructions { background-color: #F0F8FF; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #4682B4; margin-bottom: 20px; }
        .file-list { background-color: #F5F5F5; color: #333333; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 10px; }
        .stButton>button { background-color: #4CAF50; color: white; border: none; padding: 10px 20px; border-radius: 5px; font-weight: bold; }
        .stButton>button:hover { background-color: #45A049; }
        .success-box { background-color: #E6FFE6; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #2ECC71; margin-top: 20px; margin-bottom: 20px; }
        .error-box { background-color: #FFE6E6; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #FF4B4B; margin-top: 20px; margin-bottom: 20px; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="title">Excel File Merger (with Validation Summary)</div>', unsafe_allow_html=True)
    st.sidebar.header("⚙️ Diff Color Thresholds")
    low_threshold = st.sidebar.number_input("Green Threshold (≤)", min_value=0.0, max_value=1.0, value=0.1, step=0.01, key="mrg_low_threshold_sidebar")
    mid_threshold = st.sidebar.number_input("Amber Threshold (≤)", min_value=0.0, max_value=1.0, value=0.5, step=0.01, key="mrg_mid_threshold_sidebar")

    st.markdown("""
    <div class="instructions">
    <h3 style="color: #4682B4;">How to Use:</h3>
    <ul>
        <li>Upload up to 10 Excel files (outputs from Validation Report Generator).</li>
        <li>Click the "Merge and Summarize Files" button.</li>
        <li>Sheets from each file will be merged. An "All_Pages_Summary" sheet will be added first.</li>
        <li>Duplicate sheet names get a numeric suffix (e.g., 'Sheet_1'). Sheet names are limited to 31 characters.</li>
        <li>Output file name uses the first file's prefix.</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Drop Your Validation Excel Files Here!",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Upload up to 10 validation Excel files.",
        key="mrg_file_uploader_main"
    )

    if uploaded_files:
        if len(uploaded_files) > 10:
            st.markdown(
                '<div class="error-box">Whoops! Maximum 10 files allowed. Please upload fewer files.</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(f'<div class="file-list"><strong>Uploaded {len(uploaded_files)} File(s):</strong>', unsafe_allow_html=True)
            for file_obj_display in uploaded_files:
                st.markdown(f"- {file_obj_display.name}", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            if st.button("Merge and Summarize Files", key="mrg_merge_process_button"):
                with st.spinner("Merging your files and generating summary... Hang tight!"):
                    result = combine_excel_files(uploaded_files, low_threshold, mid_threshold)
                    if result:
                        output_buffer, output_filename = result
                        st.markdown(
                            f'<div class="success-box">Success! Your merged file is ready: <strong>{output_filename}</strong></div>',
                            unsafe_allow_html=True
                        )
                        st.download_button(
                            label="Download Your Merged Excel!",
                            data=output_buffer,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="mrg_download_merged_button"
                        )

if __name__ == "__main__":
    run()
