import streamlit as st
import pandas as pd
import io
import numpy as np
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import base64  # For base64 image encoding

# Define the checklist data as a DataFrame (assuming it's used or defined elsewhere if not directly in run)
checklist_data = {
    "S.No": range(1, 8),
    "Checklist": [
        "All the columns of excel replicated in PBI (No extra columns)",
        "All the filters of excel replicated in PBI",
        "Filters working as expected (single/multi select as usual)",
        "Column names matching with excel",
        "Currency symbols to be replicated",
        "Pre-applied filters while generating validation report?",
        "Sorting is replicated"
    ],
}
checklist_df = pd.DataFrame(checklist_data)


# --- generate_validation_report function (includes "Summary Avg Diff: X.XX%" modification) ---
def generate_validation_report(excel_df, pbi_df):
    dims = [col for col in excel_df.columns if col in pbi_df.columns and
            (excel_df[col].dtype == 'object' or '_id' in col.lower() or '_key' in col.lower() or
             '_ID' in col or '_KEY' in col)]

    excel_df[dims] = excel_df[dims].fillna('NAN')
    pbi_df[dims] = pbi_df[dims].fillna('NAN')

    excel_measures = [col for col in excel_df.columns if col not in dims and np.issubdtype(excel_df[col].dtype, np.number)]
    pbi_measures = [col for col in pbi_df.columns if col not in dims and np.issubdtype(pbi_df[col].dtype, np.number)]

    all_measures = list(set(excel_measures) & set(pbi_measures))

    excel_agg = excel_df.groupby(dims, observed=False)[all_measures].sum().reset_index()
    pbi_agg = pbi_df.groupby(dims, observed=False)[all_measures].sum().reset_index()


    excel_agg['unique_key'] = excel_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()
    pbi_agg['unique_key'] = pbi_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()

    excel_agg = excel_agg[['unique_key'] + [col for col in excel_agg.columns if col != 'unique_key']]
    pbi_agg = pbi_agg[['unique_key'] + [col for col in pbi_agg.columns if col != 'unique_key']]

    # Create a DataFrame for rows that are not the summary
    unique_keys_data = list(set(excel_agg['unique_key']) | set(pbi_agg['unique_key']))
    if not unique_keys_data: # Handle case with no common keys or data after filtering
        # If there are no data rows, we still might want to show a summary (e.g., of total sums)
        # For now, we'll proceed, and it might result in an empty data section below summary.
        # Or, decide on specific behavior for fully empty data.
         pass

    data_rows_df = pd.DataFrame({'unique_key': unique_keys_data})


    for dim in dims:
        # Map from excel_agg first
        map_excel = dict(zip(excel_agg['unique_key'], excel_agg[dim]))
        data_rows_df[dim] = data_rows_df['unique_key'].map(map_excel)
        # Fill NaNs with values from pbi_agg for the same dimension
        map_pbi = dict(zip(pbi_agg['unique_key'], pbi_agg[dim]))
        data_rows_df[dim].fillna(data_rows_df['unique_key'].map(map_pbi), inplace=True)


    data_rows_df['presence'] = data_rows_df['unique_key'].apply(
        lambda key: 'Present in Both' if key in excel_agg['unique_key'].values and key in pbi_agg['unique_key'].values
        else ('Present in excel' if key in excel_agg['unique_key'].values
              else 'Present in PBI')
    )

    for measure in all_measures:
        data_rows_df[f'{measure}_excel'] = data_rows_df['unique_key'].map(dict(zip(excel_agg['unique_key'], excel_agg[measure])))
        data_rows_df[f'{measure}_PBI'] = data_rows_df['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[measure])))

        # Calculate Diff for data rows
        excel_vals = data_rows_df[f'{measure}_excel'].fillna(0)
        pbi_vals = data_rows_df[f'{measure}_PBI'].fillna(0)

        # Diff calculation logic for data rows (sum of measures for a given unique_key combination)
        # This diff is absolute difference of sums.
        # If you need percentage diff for each row: (pbi_val - excel_val) / excel_val
        # The current code calculates absolute difference.
        # The request implies the _Diff column for summary is a percentage, but for data rows it's also a percentage of row-level sums.
        # Let's assume the original _Diff for data rows was also intended as a relative percentage difference.
        # (PBI_sum_for_key - Excel_sum_for_key) / Excel_sum_for_key

        diff_values = np.where(
            (excel_vals == 0) & (pbi_vals == 0), 0, # Both zero, diff is 0
            np.where(excel_vals == 0, 1, # Excel is zero, PBI is not, 100% diff (or -1 if PBI is negative)
                     abs((pbi_vals - excel_vals) / excel_vals) # Standard percentage diff
            )
        )
        data_rows_df[f'{measure}_Diff'] = np.round(diff_values, 4)


    # Summary Row Calculation
    summary_row_data = {'unique_key': 'Summary'} # Placeholder
    for dim in dims:
        summary_row_data[dim] = ''
    summary_row_data['presence'] = '' # Placeholder for presence summary string
    for measure in all_measures:
        summary_row_data[f'{measure}_excel'] = excel_df[measure].sum() # Overall sum from original excel_df
        summary_row_data[f'{measure}_PBI'] = pbi_df[measure].sum()     # Overall sum from original pbi_df
        summary_row_data[f'{measure}_Diff'] = '' # Placeholder for overall diff percentage

    summary_row = pd.Series(summary_row_data)

    diff_percentages_for_average = []
    for measure in all_measures:
        excel_total_sum = excel_df[measure].sum()
        pbi_total_sum = pbi_df[measure].sum()
        diff_percentage = 0
        if excel_total_sum != 0:
            diff_percentage = abs(round((pbi_total_sum - excel_total_sum) / excel_total_sum, 4))
        elif pbi_total_sum != 0: # Excel sum is 0, PBI sum is not
            diff_percentage = 1  # 100% difference
        # If both are 0, diff_percentage remains 0
        summary_row[f'{measure}_Diff'] = diff_percentage
        if pd.notna(diff_percentage):
            diff_percentages_for_average.append(diff_percentage)

    avg_diff = 0
    if diff_percentages_for_average:
        avg_diff = sum(diff_percentages_for_average) / len(diff_percentages_for_average)
    summary_label = f"Avg Diff: {avg_diff * 100:.2f}%"

    # Presence counts for summary (based on data_rows_df before summary is added)
    present_in_both_count = data_rows_df['presence'].str.contains('Both', na=False).sum()
    present_in_excel_only_count = data_rows_df['presence'].eq('Present in excel').sum()
    present_in_pbi_only_count = data_rows_df['presence'].eq('Present in PBI').sum()
    summary_row['presence'] = f'Both: {present_in_both_count}, Excel: {present_in_excel_only_count}, PBI: {present_in_pbi_only_count}'


    column_order = ['unique_key'] + dims + ['presence'] + \
                   [col for measure_col in all_measures for col in
                    [f'{measure_col}_excel', f'{measure_col}_PBI', f'{measure_col}_Diff']]

    summary_df = summary_row.reindex(column_order).to_frame().T
    summary_df['unique_key'] = summary_label # Set the calculated average diff label

    # Concatenate summary with data rows
    final_validation_report = pd.concat([summary_df, data_rows_df], ignore_index=True)
    final_validation_report = final_validation_report[column_order] # Ensure final column order

    return final_validation_report, excel_agg, pbi_agg

# --- column_checklist function ---
def column_checklist(excel_df, pbi_df):
    excel_columns = excel_df.columns.tolist()
    pbi_columns = pbi_df.columns.tolist()
    max_len = max(len(excel_columns), len(pbi_columns))
    excel_cols_padded = excel_columns + [''] * (max_len - len(excel_columns))
    pbi_cols_padded = pbi_columns + [''] * (max_len - len(pbi_columns))
    checklist_df = pd.DataFrame({
        'Excel Columns': excel_cols_padded,
        'PowerBI Columns': pbi_cols_padded
    })
    checklist_df['Match'] = checklist_df.apply(lambda row: row['Excel Columns'] == row['PowerBI Columns'] if row['Excel Columns'] and row['PowerBI Columns'] else False, axis=1)
    return checklist_df

# --- generate_diff_checker function ---
def generate_diff_checker(validation_report): # validation_report here is the final one with summary row
    if validation_report.empty:
        return pd.DataFrame({
            'Diff Column Name': ['No data to check'],
            'Percentage Difference': ['N/A']
        })

    diff_columns = [col for col in validation_report.columns if col.endswith('_Diff')]
    summary_row_values = validation_report.iloc[0] # This is the summary row

    diff_checker_data = []
    for col in diff_columns:
        value = summary_row_values[col]
        if pd.notna(value) and isinstance(value, (int, float)):
            diff_checker_data.append({'Diff Column Name': col, 'Percentage Difference': f"{value * 100:.2f}%"})
        else:
             # if it's already a string from summary (like 'N/A' or pre-formatted)
            diff_checker_data.append({'Diff Column Name': col, 'Percentage Difference': str(value)})

    diff_checker = pd.DataFrame(diff_checker_data)
    presence_value_summary_row = summary_row_values['presence']

    try:
        # Parse the "Both: X, Excel: Y, PBI: Z" string
        parts = presence_value_summary_row.split(',')
        both_count_str = parts[0].split(':')[1].strip()
        excel_count_str = parts[1].split(':')[1].strip()
        # pbi_count_str = parts[2].split(':')[1].strip() # If needed

        both_count = int(both_count_str)
        excel_only_count = int(excel_count_str)

        total_for_presence_metric = both_count + excel_only_count
        presence_percentage_metric = (both_count / total_for_presence_metric * 100) if total_for_presence_metric > 0 else 0
        # Refined presence summary text
        presence_summary_text = f"{presence_percentage_metric:.2f}% ({both_count} Both / {total_for_presence_metric} Total (Both+ExcelOnly))"

    except Exception: # Fallback if parsing fails
        presence_summary_text = presence_value_summary_row if pd.notna(presence_value_summary_row) else 'N/A'


    presence_summary_df = pd.DataFrame([{
        'Diff Column Name': 'Row Presence (Both / (Both + Excel Only))',
        'Percentage Difference': presence_summary_text
    }])
    diff_checker = pd.concat([diff_checker, presence_summary_df], ignore_index=True)
    return diff_checker


def run():
    st.markdown("""
        <style>
        .title { font-size: 36px; color: #FF4B4B; text-align: center; font-weight: bold; margin-bottom: 20px; }
        .instructions { background-color: #F0F8FF; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #4682B4; margin-bottom: 20px; }
        .file-list { background-color: #F5F5F5; color: #333333; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 10px; }
        .stButton>button { background-color: #4CAF50; color: white; border: none; padding: 10px 20px; border-radius: 5px; font-weight: bold; }
        .stButton>button:hover { background-color: #45A049; }
        .success-box { background-color: #E6FFE6; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #2ECC71; margin-top: 20px; margin-bottom: 20px; }
        .error-box { background-color: #FFE6E6; color: #333333; padding: 15px; border-radius: 10px; border-left: 5px solid #FF4B4B; margin-top: 20px; margin-bottom: 20px; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="title">Validation Report Generator</div>', unsafe_allow_html=True)

    st.sidebar.header("⚙️ Diff Color Thresholds")
    low_threshold = st.sidebar.number_input("Green Threshold (≤)", min_value=0.0, max_value=1.0, value=0.1, step=0.01)
    mid_threshold = st.sidebar.number_input("Amber Threshold (≤)", min_value=0.0, max_value=1.0, value=0.5, step=0.01)

    st.markdown("""
    <div class="instructions">
    <h3 style="color: #4682B4;">How to Use:</h3>
    <ul>
        <li>Upload an Excel file with two sheets: "excel" and "PBI".</li>
        <li>Ensure column names are similar for accurate comparison.</li>
        <li>Include "_ID" or "_KEY" in ID/Key/Code column names (case insensitive).</li>
        <li>Preview and download your formatted Excel report!</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)

    uploaded_file = st.file_uploader("Drop Your Excel File Here!", type=["xls", "xlsx"], help="Upload Excel with 'excel' & 'PBI' sheets.")

    if uploaded_file is not None:
        st.markdown(f'<div class="file-list"><strong>Uploaded File:</strong> {uploaded_file.name}</div>', unsafe_allow_html=True)
        with st.spinner("Generating your validation report... Hang tight!"):
            try:
                xls = pd.ExcelFile(uploaded_file)
                excel_df_orig = pd.read_excel(xls, 'excel')
                pbi_df_orig = pd.read_excel(xls, 'PBI')

                excel_df = excel_df_orig.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)
                pbi_df = pbi_df_orig.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)

                validation_report, excel_agg, pbi_agg = generate_validation_report(excel_df.copy(), pbi_df.copy())
                column_checklist_df = column_checklist(excel_df_orig, pbi_df_orig) # Use original for checklist case sensitivity if needed
                diff_checker_df = generate_diff_checker(validation_report)

                st.subheader("Validation Report Preview")
                display_report = validation_report.copy()
                # Format _Diff columns for display
                for col_name_display in display_report.columns:
                    if col_name_display.endswith('_Diff'):
                        # The first row (summary) _Diff is already a percentage (0-1), format it.
                        # Other rows _Diff are also percentages (0-1), format them too.
                        # The unique_key of summary is text, so it won't be affected.
                        def format_diff_for_st_display(val):
                            if pd.notna(val) and isinstance(val, (int, float)):
                                return f"{val * 100:.2f}%"
                            return val # if it's already text (like the summary unique_key) or NaN
                        display_report[col_name_display] = display_report[col_name_display].apply(format_diff_for_st_display)
                st.dataframe(display_report)


                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    original_filename = os.path.splitext(uploaded_file.name)[0]
                    sheet_name_report = f"{original_filename}_validation_report"[:31]
                    validation_report.to_excel(writer, sheet_name=sheet_name_report, index=False)
                    ws_report = writer.sheets[sheet_name_report]

                    def apply_conditional_formatting(ws, report_df, low_thresh, mid_thresh):
                        # --- Color definitions ---
                        dark_green_fill = PatternFill(start_color='19D119', end_color='19D119', fill_type='solid')
                        dark_red_fill = PatternFill(start_color='E82D1C', end_color='E82D1C', fill_type='solid')
                        amber_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        
                        # Summary row fill (lighter blue)
                        summary_row_blue = '96DED1'
                        summary_fill = PatternFill(start_color=summary_row_blue, end_color=summary_row_blue, fill_type='solid')
                        
                        # Header row fill (SLIGHTLY DARKER BLUE)
                        header_row_darker_blue = '6495ED' # Darker than 80BBD9
                        header_fill = PatternFill(start_color=header_row_darker_blue, end_color=header_row_darker_blue, fill_type='solid')
                        
                        bold_font = Font(bold=True)

                        # --- Header Formatting (Row 1) ---
                        for cell in ws[1]:
                            cell.font = bold_font
                            cell.fill = header_fill # Use new darker blue for header

                        # --- Summary Row Formatting (Row 2 in Excel) ---
                        for col_idx, col_name in enumerate(report_df.columns, 1):
                            cell = ws.cell(row=2, column=col_idx)
                            cell.font = bold_font
                            cell.fill = summary_fill # Use lighter blue for summary row
                            if col_name.endswith('_Diff') and isinstance(report_df.iloc[0][col_name], (float, int)): # Check if value is numeric
                                cell.number_format = '0.00%'
                            # The 'unique_key' for summary is text: "Avg Diff: X.XX%"
                            # The 'presence' for summary is text: "Both: X, Excel: Y, PBI: Z"

                        # --- Data Rows Formatting (From Row 3 in Excel) ---
                        presence_col_letter = get_column_letter(report_df.columns.get_loc('presence') + 1)
                        for row_idx_df in range(1, len(report_df)): # DF index 1 is Excel row 3
                            excel_row_num = row_idx_df + 2
                            presence_cell_val = report_df.loc[row_idx_df, 'presence']
                            cell_to_format_presence = ws[f'{presence_col_letter}{excel_row_num}']
                            if presence_cell_val == 'Present in Both':
                                cell_to_format_presence.fill = dark_green_fill
                            elif presence_cell_val in ['Present in excel', 'Present in PBI']:
                                cell_to_format_presence.fill = dark_red_fill

                            for col_idx_data, col_name_data in enumerate(report_df.columns, 1):
                                if col_name_data.endswith('_Diff'):
                                    value = report_df.loc[row_idx_df, col_name_data]
                                    cell_to_format_diff = ws.cell(row=excel_row_num, column=col_idx_data)
                                    if pd.notna(value) and isinstance(value, (float, int)):
                                        cell_to_format_diff.number_format = '0.00%'
                                        if value <= low_thresh: cell_to_format_diff.fill = dark_green_fill
                                        elif value <= mid_thresh: cell_to_format_diff.fill = amber_fill
                                        else: cell_to_format_diff.fill = dark_red_fill
                        
                        for col_idx, column_name in enumerate(report_df.columns, 1):
                            column_letter = get_column_letter(col_idx)
                            max_length = len(str(column_name)) # Start with header length
                            for cell_val_series in report_df[column_name].astype(str):
                                if len(cell_val_series) > max_length: max_length = len(cell_val_series)
                            adjusted_width = (max_length + 2) if max_length > 0 else len(str(column_name)) + 5
                            ws.column_dimensions[column_letter].width = min(adjusted_width, 45)

                    apply_conditional_formatting(ws_report, validation_report, low_threshold, mid_threshold)

                    # Create Column_Checklist sheet
                    sheet_name_checklist = "Column_Checklist"[:31]
                    column_checklist_df.to_excel(writer, sheet_name=sheet_name_checklist, index=False)
                    ws_checklist = writer.sheets[sheet_name_checklist]
                    match_col_letter = get_column_letter(column_checklist_df.columns.get_loc('Match') + 1)
                    light_green_excel = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    light_red_excel = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    for row_num in range(2, len(column_checklist_df) + 2):
                        cell = ws_checklist[f'{match_col_letter}{row_num}']
                        if cell.value == True: cell.fill = light_green_excel
                        elif cell.value == False: cell.fill = light_red_excel
                    for col_idx, column_name in enumerate(column_checklist_df.columns, 1):
                        column_letter = get_column_letter(col_idx)
                        max_col_len = max((column_checklist_df[column_name].astype(str).map(len).max(skipna=True)), len(str(column_name)))
                        if pd.isna(max_col_len): max_col_len = len(str(column_name))
                        ws_checklist.column_dimensions[column_letter].width = min(int(max_col_len) + 2, 40)
                    ws_checklist.sheet_state = 'hidden' # HIDE THE SHEET

                    # Create Diff_Checker_Summary sheet
                    sheet_name_diff_checker = "Diff_Checker_Summary"[:31]
                    diff_checker_df.to_excel(writer, sheet_name=sheet_name_diff_checker, index=False)
                    ws_diff_checker = writer.sheets[sheet_name_diff_checker]
                    for col_idx, column_name in enumerate(diff_checker_df.columns, 1):
                        column_letter = get_column_letter(col_idx)
                        max_col_len = max((diff_checker_df[column_name].astype(str).map(len).max(skipna=True)), len(str(column_name)))
                        if pd.isna(max_col_len): max_col_len = len(str(column_name))
                        ws_diff_checker.column_dimensions[column_letter].width = min(int(max_col_len) + 2, 50)
                    ws_diff_checker.sheet_state = 'hidden' # HIDE THE SHEET


                output.seek(0)
                new_file_name = f"{original_filename}_validation_report.xlsx"
                st.markdown(f'<div class="success-box">Success! Your validation report is ready: <strong>{new_file_name}</strong></div>', unsafe_allow_html=True)
                st.download_button("Download Your Validation Report!", output, new_file_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"An error occurred during report generation: {e}")
                import traceback
                st.error(traceback.format_exc())
    st.markdown("---")

if __name__ == "__main__":
    run()